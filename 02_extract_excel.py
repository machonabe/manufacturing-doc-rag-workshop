# Databricks notebook source
# DBTITLE 1,02 Excel抽出: 紹介
# MAGIC %md
# MAGIC # 02_extract_excel: Excel 試験成績書からのメタデータ抽出
# MAGIC
# MAGIC ## 目的
# MAGIC - Excel ファイルから埋め込み画像（波形チャート）を分離抽出
# MAGIC - 取り消し線（修正履歴）の検出とメタデータ化
# MAGIC - 波形生データの CSV 保存
# MAGIC - ファイル名・セル内容から製品ID / 仕様書ID / 試験ID を抽出
# MAGIC
# MAGIC ## 所要時間目安: 20分
# MAGIC
# MAGIC ## 前提
# MAGIC - 01 ノートブックが実行済み（raw/ に Excel ファイルが存在）
# MAGIC
# MAGIC ## このノートブックで作成されるもの
# MAGIC - `media_assets` テーブル（図・波形とファイルの紐付け）
# MAGIC - `excel_cells` テーブル（セル単位メタデータ）
# MAGIC - `documents` テーブル（Excel分）
# MAGIC - `images/` に抽出された PNG ファイル
# MAGIC - `waveforms/` に抽出された CSV ファイル
# MAGIC
# MAGIC ## フォーマット不統一への現実的な向き合い方
# MAGIC 製造業の試験成績書は、部署・時期・担当者によってフォーマットが異なります。
# MAGIC 完璧な構造化よりも、**「どの製品・どの仕様・どの試験に紐付くか」というメタデータの抽出を最優先**します。
# MAGIC これにより、後段の検索で「関連する図・波形・仕様書」を即座に表示できます。

# COMMAND ----------

# DBTITLE 1,ライブラリ・設定
# MAGIC %pip install openpyxl --quiet

# COMMAND ----------

# DBTITLE 1,設定読み込み
# MAGIC %run ./00_config

# COMMAND ----------

# DBTITLE 1,Excel抽出メインロジック
# ==============================================================
# Excel ファイルの走査とメタデータ抽出
# ==============================================================
import openpyxl
from openpyxl.utils import get_column_letter
import os, re, csv, zipfile, uuid
from datetime import datetime
from io import BytesIO
from pyspark.sql.types import *
from pyspark.sql import Row

xlsx_files = sorted([f for f in os.listdir(RAW_DIR) if f.endswith('.xlsx')])
print(f"処理対象 Excel ファイル: {len(xlsx_files)} 件")

# 抽出結果を貯めるリスト
doc_records = []      # documents テーブル用
media_records = []    # media_assets テーブル用
cell_records = []     # excel_cells テーブル用

def extract_ids_from_text(text):
    """テキストから製品ID/仕様書ID/試験IDを正規表現で抽出"""
    product_ids = list(set(re.findall(r'SNS-\d{3}', str(text))))
    spec_ids = list(set(re.findall(r'SPEC-SNS-\d{3}-v\d+', str(text))))
    test_ids = list(set(re.findall(r'TEST-\d{4}-\d{3}', str(text))))
    return product_ids, spec_ids, test_ids

def extract_images_openpyxl(ws, file_name, sheet_name):
    """方式1: openpyxl APIで埋め込み画像を抽出"""
    results = []
    for img in ws._images:
        anchor = img.anchor
        if hasattr(anchor, '_from'):
            col = get_column_letter(anchor._from.col + 1)
            row_num = anchor._from.row + 1
            cell_ref = f"{col}{row_num}"
        else:
            cell_ref = "unknown"
        img_data = img._data()
        asset_id = str(uuid.uuid4())[:8]
        img_fname = f"{file_name}_{sheet_name}_{cell_ref}_{asset_id}.png"
        img_path = os.path.join(IMAGES_DIR, img_fname)
        with open(img_path, 'wb') as f:
            f.write(img_data)
        results.append({
            "asset_id": asset_id,
            "file_path": img_path,
            "source_sheet": sheet_name,
            "anchor_cell": cell_ref,
            "size_bytes": len(img_data)
        })
    return results

def extract_images_zipfile(filepath, file_name):
    """方式2: zipfileフォールバック（xl/media/から直接抽出）"""
    results = []
    try:
        with zipfile.ZipFile(filepath, 'r') as zf:
            media_files = [n for n in zf.namelist() if n.startswith('xl/media/')]
            for mf in media_files:
                data = zf.read(mf)
                ext = os.path.splitext(mf)[1]
                asset_id = str(uuid.uuid4())[:8]
                img_fname = f"{file_name}_zip_{asset_id}{ext}"
                img_path = os.path.join(IMAGES_DIR, img_fname)
                with open(img_path, 'wb') as f:
                    f.write(data)
                results.append({"asset_id": asset_id, "file_path": img_path,
                               "source_sheet": "zip_extract", "anchor_cell": ""})
    except Exception as e:
        print(f"  ⚠️ zipfile抽出失敗: {e}")
    return results

# --- メインループ: 各Excelファイルを処理 ---
for fname in xlsx_files:
    fpath = os.path.join(RAW_DIR, fname)
    print(f"\n処理中: {fname}")
    
    # ファイル名からID抽出
    prod_ids, spec_ids, test_ids = extract_ids_from_text(fname)
    all_text = fname  # 全テキストを累積（後でID抽出に使用）
    
    wb = openpyxl.load_workbook(fpath)
    file_images = []
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        
        # (a) 埋め込み画像抽出
        imgs = extract_images_openpyxl(ws, fname.replace('.xlsx',''), ws_name)
        if not imgs:
            imgs = extract_images_zipfile(fpath, fname.replace('.xlsx',''))
        file_images.extend(imgs)
        
        # (b) セル走査: 値・取り消し線・結合セル
        merged = set()
        for mr in ws.merged_cells.ranges:
            for cell in mr.cells:
                merged.add(cell)
        
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 50)):
            for cell in row:
                if cell.value is not None:
                    all_text += f" {cell.value}"
                    is_strike = bool(cell.font and cell.font.strike)
                    is_merged = (cell.row, cell.column) in merged
                    cell_records.append(Row(
                        file_name=fname, sheet=ws_name,
                        cell=f"{get_column_letter(cell.column)}{cell.row}",
                        value=str(cell.value)[:500],
                        is_strikethrough=is_strike, is_merged=is_merged
                    ))
        
        # (c) 波形データ検出・保存
        for row in ws.iter_rows(min_row=1, max_row=5, max_col=5):
            for cell in row:
                if cell.value and "time" in str(cell.value).lower():
                    # 波形データシートをCSVに保存
                    wave_fname = f"EXTRACTED_{fname.replace('.xlsx','')}_{ws_name}.csv"
                    wave_path = os.path.join(WAVEFORMS_DIR, wave_fname)
                    data_rows = []
                    header_row = cell.row
                    for r in ws.iter_rows(min_row=header_row, max_col=2):
                        vals = [c.value for c in r]
                        if all(v is not None for v in vals):
                            data_rows.append(vals)
                    if len(data_rows) > 1:
                        with open(wave_path, 'w', newline='') as f:
                            writer = csv.writer(f)
                            for dr in data_rows:
                                writer.writerow(dr)
                        media_records.append(Row(
                            asset_id=str(uuid.uuid4())[:8],
                            asset_type="waveform_csv",
                            file_path=wave_path,
                            thumbnail_path="",
                            source_doc_id=fname,
                            source_file_name=fname,
                            source_sheet=ws_name,
                            anchor_cell="",
                            product_id=prod_ids[0] if prod_ids else "",
                            spec_id=spec_ids[0] if spec_ids else "",
                            test_id=test_ids[0] if test_ids else "",
                            description=f"波形データ({ws_name}) - {fname}"
                        ))
                    break
    
    # 画像を media_assets に登録
    for img_info in file_images:
        media_records.append(Row(
            asset_id=img_info['asset_id'],
            asset_type="waveform_chart",
            file_path=img_info['file_path'],
            thumbnail_path="",
            source_doc_id=fname,
            source_file_name=fname,
            source_sheet=img_info['source_sheet'],
            anchor_cell=img_info['anchor_cell'],
            product_id=prod_ids[0] if prod_ids else "",
            spec_id=spec_ids[0] if spec_ids else "",
            test_id=test_ids[0] if test_ids else "",
            description=f"埋め込みチャート({img_info['source_sheet']}!{img_info['anchor_cell']}) - {fname}"
        ))
    
    # 全テキストからID再抽出（セル内容も含む）
    all_prod_ids, all_spec_ids, all_test_ids = extract_ids_from_text(all_text)
    
    doc_records.append(Row(
        doc_id=fname, file_name=fname,
        file_path=os.path.join(RAW_DIR, fname),
        doc_type="excel",
        title=f"Excel試験成績書 - {fname}",
        product_ids=all_prod_ids,
        spec_ids=all_spec_ids,
        keywords=[],
        summary="",
        ingested_at=datetime.now().isoformat()
    ))
    
    wb.close()
    print(f"  → 画像: {len(file_images)}, 取消線: {sum(1 for c in cell_records if c.file_name==fname and c.is_strikethrough)}")

print(f"\n{'='*50}")
print(f"抽出完了:")
print(f"  documents: {len(doc_records)} 件")
print(f"  media_assets: {len(media_records)} 件")
print(f"  excel_cells: {len(cell_records)} 件")

# COMMAND ----------

# DBTITLE 1,Deltaテーブルへの書き込み
# ==============================================================
# Delta テーブルへの書き込み
# ==============================================================
# media_assets
if media_records:
    df_media = spark.createDataFrame(media_records)
    df_media.write.mode("overwrite").saveAsTable(FQ_MEDIA_ASSETS)
    print(f"✅ {FQ_MEDIA_ASSETS}: {df_media.count()} 件")

# excel_cells
if cell_records:
    df_cells = spark.createDataFrame(cell_records)
    df_cells.write.mode("overwrite").saveAsTable(FQ_EXCEL_CELLS)
    print(f"✅ {FQ_EXCEL_CELLS}: {df_cells.count()} 件")

# documents (Excel分)
if doc_records:
    df_docs = spark.createDataFrame(doc_records)
    df_docs.write.mode("overwrite").saveAsTable(FQ_DOCUMENTS)
    print(f"✅ {FQ_DOCUMENTS}: {df_docs.count()} 件")

# COMMAND ----------

# DBTITLE 1,検証
# ==============================================================
# 検証
# ==============================================================
# 抽出画像の確認
extracted_images = [f for f in os.listdir(IMAGES_DIR) if f.endswith('.png')]
assert len(extracted_images) >= 4, f"抽出画像数不足: {len(extracted_images)}"
print(f"抽出画像: {len(extracted_images)} 件")

# 取り消し線検出の確認
strike_count = spark.sql(f"SELECT COUNT(*) as cnt FROM {FQ_EXCEL_CELLS} WHERE is_strikethrough = true").first().cnt
assert strike_count >= 4, f"取り消し線検出数不足: {strike_count}"
print(f"取り消し線セル: {strike_count} 件")

# サンプル表示: 取り消し線セルの一覧
print("\n取り消し線が検出されたセル:")
spark.sql(f"""
  SELECT file_name, sheet, cell, value 
  FROM {FQ_EXCEL_CELLS} 
  WHERE is_strikethrough = true
""").show(truncate=False)

print("✅ 02_extract_excel 完了")
